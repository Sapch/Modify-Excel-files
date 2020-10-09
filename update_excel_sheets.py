"""
This module could be used to automate creation and modification of Excel sheets
specifically is useful when you are dealing with

Working functions:
    - updating values of specific column in all sheet
    - updating column titles of all the sheets
    - Creating Excel sheet/s with its column headers

"""
import openpyxl as xl


def update_col_values(name: "Excel file name", col: int, correction: int):
    """ Modifies all values of a specific column in Excel sheets
    Usage: updating excel file containing several sheets that are similar
    and all sheets are needed to updated similarly

    Parameters:
        name -- Excel file name
        col -- col number to be corrected
        correction -- a correction value between 0 to 1
    """
    print(f"*** Excel file: {name} ***")
    wb = xl.load_workbook(name)
    for sheet in wb.worksheets:
        print(f"working on {sheet}")
        for row in range(2, sheet.max_row + 1):
            current_value = sheet.cell(row, col).value
            corrected_value = current_value * correction
            sheet.cell(row, col).value = corrected_value
            print(f"row{row}: {current_value} --> {corrected_value}")

    wb.save(name)


def update_col_headers(name: "Excel file name", new_headings: list):
    """ Modifies headers of columns in all of the sheets
    Usage: updating excel file containing several sheets that are similar
    and all sheets are needed to updated similarly

    Parameters:
        name -- Excel file name
        new_headings -- list of new titles
    """
    print(f"*** Excel file: {name} ***")
    wb = xl.load_workbook(name)
    for sheet in wb.worksheets:
        print(f"working on {sheet}")
        for col in range(1, len(new_headings)+1):
            current_title = sheet.cell(1, col).value
            updated_title = new_headings[col-1]
            sheet.cell(1, col).value = updated_title
            print(f"{current_title} --> {updated_title}")

    wb.save(name)


def creat_excel_sheet(name: "Excel file name", col_headings: list, sheet_names: list):
    """Creates Excel sheet/s with its column headers
    Usage: creating several sheets within an Excel file with similar column headings

    name -- name of the excel file to be created
    col_headings -- Excel sheets column headers
    sheet_name -- names of the sheets
    """
    wb = xl.Workbook()
    for sheet in sheet_names:
        wb.create_sheet(sheet, 0)
        sheet = wb[sheet]

        col = 1  # first col in the sheet
        for header in col_headings:
            sheet.cell(1, col).value = header
            col += 1

    wb.save(name+".xlsx")